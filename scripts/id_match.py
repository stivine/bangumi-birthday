"""
Excel 批量 ID 匹配工具

从 character.jsonlines 构建反向索引（名称 + 生日 → character_id），
然后对 Excel 文件中每行的"人物名+生日"进行匹配，将找到的 ID 写入 A 列。

用法：
    python scripts/id_match.py \
        --excel /path/to/相关人物.xlsm \
        --output /path/to/相关人物_匹配后.xlsm \
        [--overwrite]  # 是否覆盖 A 列已有值

依赖：
    pip install openpyxl

配置：
    通过 .env 文件的 BGM_DATA_DIR 指定数据目录；
    或直接在命令行传入 --data-dir。
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


# ── 将项目根路径加入 sys.path，以便导入 bangumi_birthday 包 ──────────────────
_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_ROOT))

from bangumi_birthday.utils.date_utils import extract_month_day, parse_infobox_names
from bangumi_birthday.utils.jsonlines import iter_jsonlines_with_progress


def build_index(
    jsonlines_path: Path,
) -> dict[tuple[str, str | None], set[int]]:
    """
    扫描 character.jsonlines，构建 (name, birthday) → {character_id} 的反向索引。

    Parameters
    ----------
    jsonlines_path : Path
        character.jsonlines 文件路径。

    Returns
    -------
    dict
        键为 (name_str, "MM-DD" | None)，值为 character_id 集合。
    """
    index: dict[tuple[str, str | None], set[int]] = defaultdict(set)

    for item in iter_jsonlines_with_progress(jsonlines_path, desc="构建反向索引"):
        char_id: int = int(item.get("id", 0))
        infobox: str = item.get("infobox", "")

        parsed = parse_infobox_names(infobox)
        birthday_raw = str(parsed.get("birthday_raw", ""))
        birthday_md = extract_month_day(birthday_raw) if birthday_raw else None

        # 收集所有名称
        names: list[str] = []
        top_name = item.get("name", "")
        if top_name:
            names.append(top_name.strip())
        for field in ("simplified_cn", "second_cn", "japanese"):
            val = parsed.get(field, "")
            if val and isinstance(val, str) and val.strip():
                names.append(val.strip())
        aliases = parsed.get("aliases", [])
        if isinstance(aliases, list):
            names.extend(a.strip() for a in aliases if isinstance(a, str) and a.strip())

        # 写入索引
        for name in set(names):  # 同一角色的不同名称
            index[(name, birthday_md)].add(char_id)

    logger.info("索引构建完成，共 %d 个键", len(index))
    return index


def match_excel(
    excel_path: Path,
    output_path: Path,
    index: dict[tuple[str, str | None], set[int]],
    *,
    overwrite: bool = False,
) -> dict[str, int]:
    """
    遍历 Excel，对每行进行名称+生日匹配，将 character_id 写入 A 列。

    Parameters
    ----------
    excel_path : Path
        输入 Excel 文件（支持 .xlsm，保留 VBA）。
    output_path : Path
        输出文件路径。
    index : dict
        由 build_index() 构建的反向索引。
    overwrite : bool
        若 True，覆盖 A 列已有值；否则跳过。

    Returns
    -------
    dict[str, int]
        {"matched": N, "skipped": N, "not_found": N}
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("请先安装 openpyxl：pip install openpyxl") from exc

    wb = load_workbook(excel_path, keep_vba=True)
    ws = wb.active

    stats = {"matched": 0, "skipped": 0, "not_found": 0}
    total_rows = ws.max_row

    logger.info("开始匹配 Excel（共 %d 行）...", total_rows - 1)

    for row_idx in range(2, total_rows + 1):
        # B 列：人物名；J 列：生日
        b_val = ws.cell(row=row_idx, column=2).value
        j_val = ws.cell(row=row_idx, column=10).value

        if not b_val:
            continue

        # 支持 "&" 分隔的多个名称（如"葵日向 & 葵梦月"）
        names_to_match = [n.strip() for n in str(b_val).split("&") if n.strip()]
        table_birthday = extract_month_day(str(j_val)) if j_val else None

        matched_ids: set[int] = set()
        for name in names_to_match:
            found = index.get((name, table_birthday), set())
            matched_ids.update(found)

        if not matched_ids:
            logger.debug("第 %d 行未匹配：%s / %s", row_idx, names_to_match, table_birthday)
            stats["not_found"] += 1
            continue

        a_cell = ws.cell(row=row_idx, column=1)
        a_val = a_cell.value
        is_empty = a_val is None or (isinstance(a_val, str) and not a_val.strip())

        if not overwrite and not is_empty:
            stats["skipped"] += 1
            continue

        result_str = ",".join(map(str, sorted(matched_ids)))
        a_cell.value = result_str
        stats["matched"] += 1
        logger.debug("第 %d 行 → %s", row_idx, result_str)

    wb.save(output_path)
    logger.info("已保存结果：%s", output_path)
    return stats


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S",
    )

    parser = argparse.ArgumentParser(
        description="将 Excel 中的角色名匹配为 Bangumi character ID，写入 A 列"
    )
    parser.add_argument("--excel", required=True, type=Path, help="输入 Excel 文件路径")
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="输出文件路径（默认：原文件名加 _matched 后缀）",
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=None,
        help="JSONLINES 数据目录（默认读取 BGM_DATA_DIR 环境变量）",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="覆盖 A 列已有值（默认只填充空单元格）",
    )

    args = parser.parse_args()

    # 确定数据目录
    if args.data_dir:
        jsonlines_path = args.data_dir / "character.jsonlines"
    else:
        from bangumi_birthday.config import get_settings
        jsonlines_path = get_settings().character_file

    if not jsonlines_path.exists():
        logger.error("找不到数据文件：%s", jsonlines_path)
        sys.exit(1)

    # 确定输出路径
    output_path: Path = args.output or args.excel.parent / (
        args.excel.stem + "_matched" + args.excel.suffix
    )

    # 构建索引
    logger.info("数据文件：%s", jsonlines_path)
    index = build_index(jsonlines_path)

    # 匹配
    stats = match_excel(args.excel, output_path, index, overwrite=args.overwrite)
    logger.info(
        "匹配完成：成功写入 %d 行，跳过（已有值）%d 行，未匹配 %d 行",
        stats["matched"],
        stats["skipped"],
        stats["not_found"],
    )


if __name__ == "__main__":
    main()
